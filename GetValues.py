import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Load input workbook and declare worksheet variables

inputWb = load_workbook('Unit Fcst Consolidated.xlsx', data_only=True)
workSheets = ['VN', 'SG', 'TH', 'TW', 'KR', 'KZ']
wsVN = inputWb['VN']  # VIETNAM
wsSG = inputWb['SG']  # SINGAPORE
wsTH = inputWb['TH']  # THAILAND
wsTW = inputWb['TW']  # TAIWAN
wsKR = inputWb['KR']  # KOREA
wsKZ = inputWb['KZ']  # KAZAKHSTAN

output = Workbook()

PATH = './Raw.xlsx'
if os.path.isfile(PATH) and os.access(PATH, os.R_OK):
    output = load_workbook('Raw.xlsx')
    # print("File exists and is readable")

# Load output workbook and declare worksheet variable

if 'Values' in output.sheetnames:
    output.remove(output['Values'])
output.create_sheet("Values")
ws = output['Values']

output.active = output['Values']

# Define the starting point of the range of interest
startingIndex = 'E2'  # MODIFY AS NEEDED

for sheet in workSheets:

    currSh = inputWb[sheet]
    col_num = currSh[startingIndex].col_idx
    row_num = currSh[startingIndex].row

    lastRow = ws.max_row

    # Get the maximum number of rows & columns in the working range
    max_row = len([cell for cell in currSh['A'] if cell.value])  # ROWS
    max_col = len(currSh[1])  # COLUMNS
    max_cell = currSh.cell(row=max_row, column=max_col)

    # Declare the working range from defined start & end parameters
    rg = currSh[startingIndex + ':' + str(max_cell.coordinate)]

    for i, row in enumerate(rg):
        for j, cell in enumerate(row):

            # Define an index variable that ensures values are iterated over sets of column counts
            indexManipulator = i * (len(row) - 1) + lastRow

            # Extract the SKU ID
            ws.cell(row=i + 1 + indexManipulator + j, column=1).value = currSh.cell(row=cell.row, column=1).value

            # Extract the year
            dttm = currSh.cell(row=1, column=5 + j).value
            year = dttm.year
            yearCell = ws.cell(row=i + 1 + indexManipulator + j, column=2)
            yearCell.number_format = 'General'
            yearCell.value = year

            # Extract the month
            month = dttm.month
            monthCell = ws.cell(row=i + 1 + indexManipulator + j, column=3)
            monthCell.number_format = 'General'
            monthCell.value = month

            # Copy the Channel
            ws.cell(row=i + 1 + indexManipulator + j, column=4).value = currSh.cell(row=cell.row, column=4).value

            # Copy the unit count for the corresponding month, year, channel, and SKU
            ws.cell(row=i + 1 + indexManipulator + j, column=5).value = currSh.cell(row=cell.row, column=cell.col_idx).value

ws['A1'] = "SKU"
ws['B1'] = "YEAR"
ws['C1'] = "MONTH"
ws['D1'] = "CHANNEL"
ws['E1'] = "UNITS"

max_rows = ws.max_row
max_col = ws.max_column

tab = Table(displayName="Consolidated_Forecast", ref=f"A1:{get_column_letter(max_col)}{max_rows}")

# Add a default style with striped rows and banded columns
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
tab.tableStyleInfo = style
ws.add_table(tab)

if 'Sheet' in output.sheetnames:
    output.remove(output['Sheet'])

output.save('Raw.xlsx')
